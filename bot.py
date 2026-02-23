# ============================================================
# BOT DE ANLISE DE CRDITO - V3 ULTRA HBRIDO
# API OPENAI (2025) - OCR LOCAL + FALLBACK VISION (CHAT API)
# ============================================================

import os
import sys
import json
import base64
import shutil
import tempfile
from io import BytesIO
from typing import List, Dict, Any

from PIL import Image, ImageDraw, ImageFont
from pdf2image import convert_from_path
import pytesseract

# ============================================================
# CONFIGURAÇÃO EXPLÍCITA DO TESSERACT (OBRIGATÓRIA NO WINDOWS)
# ============================================================

TESSERACT_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Tesseract-OCR\tesseract.exe"
]

TESSERACT_CMD = None
for p in TESSERACT_PATHS:
    if os.path.isfile(p):
        TESSERACT_CMD = p
        break

if not TESSERACT_CMD:
    raise RuntimeError("Tesseract OCR não encontrado. Instale e verifique o caminho.")

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
print(f"[INFO] Tesseract configurado: {TESSERACT_CMD}")

from openpyxl import load_workbook
from openpyxl.styles import Font
from openai import OpenAI
import requests
from dotenv import load_dotenv
import re

# ------------------------------------------------------------
# BASE DE BANCOS (INTERNA + OPCIONAL EXTERNA)
# ------------------------------------------------------------

LISTA_BANCOS = set()

# Palavras-chave bancárias (compacta, cobre BACEN inteiro)
BANCOS_KEYWORDS = {
    "banco", "bank", "credito", "credit", "financeira", "financiamento",
    "itau", "unibanco", "bradesco", "santander", "caixa", "bb",
    "bmg", "bs2", "btg", "safra", "pine", "daycoval", "sofisa",
    "sicredi", "sicoob", "unicred", "cresol",
    "abc", "fibra", "modal", "genial", "xp", "orama",
    "nubank", "c6", "inter", "pan", "original", "digio",
    "mercantil", "banrisul", "banestes", "bari", "votorantim",
    "volkswagen", "toyota", "ford", "honda", "gm", "mercedes",
    "jpmorgan", "citibank", "citi", "hsbc", "ubs", "credit suisse",
    "deutsche", "bnp", "paribas", "barclays", "rabobank", "ing",
    "pagseguro", "mercado pago", "cielo", "stone", "picpay",
    "leasing", "capital", "investimentos"
}

def carregar_bancos():
    global LISTA_BANCOS

    try:
        # Tenta carregar lista externa (se existir)
        path = os.path.join(EXEC_DIR, "bancos.txt")

        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                for linha in f:
                    nome = linha.strip()
                    if nome:
                        LISTA_BANCOS.add(nome.lower())

            print(f"[INFO] bancos.txt carregado: {len(LISTA_BANCOS)} entradas")

        else:
            # Fallback automático (EXE-safe)
            LISTA_BANCOS.update(BANCOS_KEYWORDS)
            print("[INFO] bancos.txt não encontrado — usando base bancária interna")

    except Exception as e:
        # Último fallback de segurança
        LISTA_BANCOS.update(BANCOS_KEYWORDS)
        print(f"[ERRO] Falha ao carregar bancos externos — fallback interno ativado: {e}")

def eh_banco(nome: str) -> bool:
    if not nome:
        return False

    nome = nome.lower()

    # Checagem por aproximação (inalterada)
    return any(b in nome for b in LISTA_BANCOS)

# ============================================================
# 1. CARREGAMENTO DO .ENV (FUNCIONA EM .PY E EM .EXE)
# ============================================================

if getattr(sys, "frozen", False):
    EXEC_DIR = os.path.dirname(sys.executable)
else:
    EXEC_DIR = os.path.dirname(os.path.abspath(__file__))

ENV_PATH = os.path.join(EXEC_DIR, ".env")

if os.path.isfile(ENV_PATH):
    print(f"[INFO] .env encontrado em: {ENV_PATH}")
    load_dotenv(ENV_PATH)
else:
    print(f"[ERRO] .env NO encontrado no diretrio: {ENV_PATH}")
    sys.exit(1)

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()

if not OPENAI_API_KEY:
    print("[ERRO] OPENAI_API_KEY no localizada no .env")
    sys.exit(1)

client = OpenAI(api_key=OPENAI_API_KEY)
print("[INFO] OpenAI client inicializado com sucesso.")


# ============================================================
# 2. POPPLER PARA PDF -> PNG
# ============================================================

DEFAULT_POPPLER_PATHS = [
    r"C:\Program Files\poppler-25.11.0\Library\bin",
    r"C:\Program Files\poppler-24.08.0\Library\bin",
    r"C:\Program Files\poppler-23.11.0\Library\bin"
]

POPPLER_PATH = None
for p in DEFAULT_POPPLER_PATHS:
    if os.path.isdir(p):
        POPPLER_PATH = p
        break

if POPPLER_PATH:
    print(f"[INFO] Poppler detectado: {POPPLER_PATH}")
else:
    print("[AVISO] Poppler NO localizado automaticamente. PDFs podem falhar.")

# ============================================================
# 3. UTILITRIOS DE PDF, OCR E IMAGEM
# ============================================================

MAX_PAGES_CONTRATO = 15
MAX_PAGES_FATURAMENTO = 5
MAX_PAGES_ENDIVIDAMENTO = 5


def pdf_to_images(pdf_path: str, max_pages: int, dpi=170) -> List[Image.Image]:
    """Converte PDF em lista de pginas (imagens PIL)."""
    try:
        if POPPLER_PATH:
            pages = convert_from_path(pdf_path, dpi=dpi, poppler_path=POPPLER_PATH)
        else:
            pages = convert_from_path(pdf_path, dpi=dpi)
    except Exception as e:
        print(f"[ERRO] Falha convertendo PDF '{pdf_path}': {e}")
        return []

    return pages[:max_pages]


def ocr_extract_text(image_list: List[Image.Image]) -> str:
    """OCR Tesseract local para extrair textos."""
    full = []
    try:
        for img in image_list:
            t = pytesseract.image_to_string(img, lang="por")
            if t.strip():
                full.append(t)
    except Exception as e:
        print(f"[ERRO OCR] {e}")

    return "\n".join(full).strip()


def pil_to_base64(img: Image.Image) -> str:
    """Converte imagem PIL -> base64 compatvel com GPT Vision."""
    buf = BytesIO()
    img.save(buf, format="JPEG")
    enc = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{enc}"


def chat_to_text(resp) -> str:
    """
    Extrai texto da resposta da API Chat 2025.
    Pode ser string ou lista de partes.
    """
    if not resp:
        return ""

    try:
        msg = resp.choices[0].message
    except:
        return ""

    content = getattr(msg, "content", "")

    if isinstance(content, str):
        return content.strip()

    if isinstance(content, list):
        texts = []
        for part in content:
            if hasattr(part, "text") and part.text:
                texts.append(part.text)
            elif isinstance(part, dict) and part.get("type") == "text":
                texts.append(part.get("text", ""))
        return "\n".join(texts).strip()

    return ""


def safe_json_loads(raw: str) -> Any:
    """Carrega JSON mesmo quando o GPT inclui texto extra fora das chaves."""
    if not raw or not isinstance(raw, str):
        return None

    raw = raw.strip()

    try:
        return json.loads(raw)
    except:
        pass

    try:
        i = raw.find("{")
        j = raw.rfind("}")
        if i != -1 and j != -1:
            return json.loads(raw[i:j+1])
    except:
        pass

    print("[ERRO] JSON invlido retornado pelo GPT.")
    return None


# ============================================================
# 4. GPT TEXTO  CHAT API 2025
# ============================================================

def run_gpt_text(prompt: str, text: str) -> str:
    """GPT TEXTO (modelo gpt-4o-mini) usando OCR como entrada."""
    try:
        full = (
            prompt.strip()
            + "\n\n--- TEXTO OCR EXTRADO ---\n\n"
            + text.strip()
        )

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{
                "role": "user",
                "content": full
            }],
            temperature=0,
            max_tokens=4000
        )

        return chat_to_text(resp)

    except Exception as e:
        print(f"[ERRO GPT TEXTO] {e}")
        return ""


# ============================================================
# 5. GPT VISION  FALLBACK MULTIMODAL
# ============================================================

def force_vision_extract(images: List[Image.Image], prompt: str) -> str:
    """Usa GPT Vision quando OCR  insuficiente."""
    print("[INFO] OCR insuficiente -> ativando GPT VISION...")

    content = [{"type": "text", "text": prompt.strip()}]

    for img in images[:20]:
            content.append({
        "type": "image_url",
        "image_url": {
            "url": pil_to_base64(img)
        }
    })


    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{
                "role": "user",
                "content": content
            }],
            temperature=0,
            max_tokens=5000
        )
        return chat_to_text(resp)

    except Exception as e:
        print(f"[ERRO GPT VISION] {e}")
        return ""

# ============================================================
# 6. MECANISMO HBRIDO OCR -> GPT
# ============================================================

def extract_with_ocr_then_vision(
    pdf_path: str,
    ocr_min_chars: int,
    prompt_vision: str,
    max_pages: int
) -> str:

    imgs = pdf_to_images(pdf_path, max_pages=max_pages)

    if not imgs:
        print("[ERRO] Nenhuma pagina convertida do PDF.")
        return ""

    print("[INFO] Executando OCR local...")
    ocr_text = ocr_extract_text(imgs)

    if len(ocr_text) >= ocr_min_chars:
        print(f"[INFO] OCR OK ({len(ocr_text)} chars) -> GPT TEXTO")
        return run_gpt_text(prompt_vision, ocr_text)

    print(f"[INFO] OCR fraco ({len(ocr_text)} chars) -> GPT VISION")
    return force_vision_extract(imgs, prompt_vision)


# ============================================================
# 7. EXTRATORES GPT  CONTRATO / FATURAMENTO / ENDIVIDAMENTO
# ============================================================

# ------------------------------------------------------------
# CONTRATO SOCIAL (SEM data de abertura  100% compatvel)
# ------------------------------------------------------------
def extract_contrato(pdf_path: str) -> Dict[str, Any]:
    prompt = """
Voc  um analista de crdito. Extraia SOMENTE o JSON abaixo.
NO invente nada. NO adicione campos alm dos listados.

FORMATO EXATO:

{
  "razao_social": "",
  "cidade": "",
  "estado": "",
  "capital_social": "",
  "socios": [
    {"nome": "", "percentual": ""}
  ]
}

Se no houver algum item, deixe "".
"""

    raw = extract_with_ocr_then_vision(
        pdf_path=pdf_path,
        ocr_min_chars=800,
        prompt_vision=prompt,
        max_pages=MAX_PAGES_CONTRATO
    )

    data = safe_json_loads(raw)

    if not data:
        return {
            "razao_social": "",
            "cidade": "",
            "estado": "",
            "capital_social": "",
            "socios": []
        }

    # Garantia da estrutura
    data.setdefault("razao_social", "")
    data.setdefault("cidade", "")
    data.setdefault("estado", "")
    data.setdefault("capital_social", "")
    data.setdefault("socios", [])

    return data


# ------------------------------------------------------------
# FATURAMENTO  12 meses, sem total anual
# ------------------------------------------------------------
def extract_faturamento(pdf_path: str) -> Dict[str, Any]:
    """
    Extrai faturamento multi-ano e multi-ms sem depender do layout do arquivo.
    O GPT retorna todos os meses encontrados, cada um com seu respectivo ano.
    """

    prompt = """
Voc  um analista financeiro.
Extraia TODOS os valores de faturamento encontrados no documento, mesmo se houver dados de mltiplos anos.

REGRAS:
- Identifique o ANO corretamente (2023, 2024 ou 2025).
- Identifique o MS corretamente (janeiro, fevereiro, etc.).
- NO invente valores.
- Use o formato brasileiro para valores: 12345,67
- Se no tiver certeza do ano, NO invente.

FORMATO OBRIGATRIO DA RESPOSTA:

{
  "dados": [
    {
      "ano": "2024",
      "mes": "fevereiro",
      "valor": "123.456,78"
    }
  ]
}

Se no encontrar nada, devolva:

{ "dados": [] }
    """

    raw = extract_with_ocr_then_vision(
        pdf_path=pdf_path,
        ocr_min_chars=350,
        prompt_vision=prompt,
        max_pages=MAX_PAGES_FATURAMENTO
    )

    data = safe_json_loads(raw)

    if not data or "dados" not in data:
        return {"dados": []}

    return data


# ------------------------------------------------------------
# ENDIVIDAMENTO  bancos & fundos
# ------------------------------------------------------------
def extract_endividamento(pdf_path: str) -> List[Dict[str, Any]]:
    prompt = """
Voc  um analista de crdito. Extraia o endividamento linha a linha.

FORMATO EXATO:

{
  "itens": [
    {
      "credor": "",
      "tipo_credor": "",
      "modalidade": "",
      "saldo_devedor": ""
    }
  ]
}

Classificao:
- Bancos tradicionais -> "banco"
- FIDC, securitizadoras, SPE, fundos -> "fundo"
- Caso duvidoso -> "outro"

Saldo em formato brasileiro 12345,67.
"""

    raw = extract_with_ocr_then_vision(
        pdf_path=pdf_path,
        ocr_min_chars=250,
        prompt_vision=prompt,
        max_pages=MAX_PAGES_ENDIVIDAMENTO
    )

    data = safe_json_loads(raw)

    if not data or "itens" not in data:
        return []

    return data["itens"]

# ============================================================
# 8. UTILITRIOS COMPLEMENTARES (CONVERSO / IMAGENS)
# ============================================================

def to_number_br(valor: str):
    """
    Converte texto em nmero (float) no padro brasileiro.
    Exemplos aceitos:
      - "R$ 12.345,67" -> 12345.67
      - "12.345,67"    -> 12345.67
      - "12345.67"     -> 12345.67
    Se no der pra converter, retorna None (clula fica em branco).
    """
    if not isinstance(valor, str):
        return None

    v = valor.replace("R$", "").replace(" ", "").strip()
    if not v:
        return None

    # remove pontos de milhar
    v = v.replace(".", "")
    # vrgula vira decimal
    v = v.replace(",", ".")

    try:
        return float(v)
    except Exception:
        return None


def insert_image(ws, img: Image.Image, cell: str):
    """
    Salva temporariamente a imagem e insere no Excel.
    """
    from openpyxl.drawing.image import Image as XLImage

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    img.save(tmp.name, "PNG")

    ximg = XLImage(tmp.name)
    ws.add_image(ximg, cell)


# ============================================================
# 9. ORGANOGRAMA (VERSO ESTVEL E NO DISTORCIDA)
# ============================================================

def build_organograma(socios: List[Dict[str, str]], razao_social: str) -> Image.Image:
    """
    Organograma profissional para 14 scios.
    - Caixas com tamanho fixo (quebra automtica de texto)
    - Fonte ajustada para caber dentro da caixa
    - Percentual ao lado da linha vertical (Modo B1)
    - Trilho curto para scios mltiplos
    - Imagem final 900  500 (encaixe perfeito no Excel A5D22)
    """

    from PIL import Image, ImageDraw, ImageFont
    import textwrap

    # ===========================
    # TAMANHO IDEAL DA IMAGEM
    # ===========================
    W, H = 900, 500

    box_w = 260
    box_h = 90
    spacing = 40

    empresa_w = 420
    empresa_h = 100

    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    BLUE = (23, 92, 131)
    BLUE_OUT = (10, 60, 100)
    WHITE = (255, 255, 255)
    BLACK = (25, 25, 25)

    # ===========================
    # FONTES
    # ===========================
    try:
        f_big = ImageFont.truetype("arial.ttf", 26)
        f_mid = ImageFont.truetype("arial.ttf", 22)
        f_small = ImageFont.truetype("arial.ttf", 18)
        f_perc = ImageFont.truetype("arial.ttf", 20)
    except:
        f_big = f_mid = f_small = f_perc = ImageFont.load_default()

    # ===========================
    # Funes auxiliares
    # ===========================
    def wrap_text(text, max_chars):
        """Quebra texto em mltiplas linhas."""
        return "\n".join(textwrap.wrap(text, width=max_chars))

    def fit_font(text, font_big, font_mid, font_small):
        """Escolhe fonte adequada ao comprimento do texto."""
        L = len(text)
        if L <= 18:
            return font_big
        elif L <= 28:
            return font_mid
        else:
            return font_small

    def draw_centered_multiline(x1, y1, x2, y2, text, font):
        """Texto multiline centralizado dentro da caixa."""
        lines = text.split("\n")
        line_height = font.getbbox("Ag")[3] + 4
        total_h = len(lines) * line_height
        y_text = y1 + (y2 - y1 - total_h) / 2

        for line in lines:
            w, h = draw.textbbox((0, 0), line, font=font)[2:]
            x_text = x1 + (x2 - x1 - w) / 2
            draw.text((x_text, y_text), line, fill=WHITE, font=font)
            y_text += line_height

    def abreviar(nome: str):
        if not nome:
            return "EMPRESA"
        return " ".join(nome.split()[:3]).upper()

    def arrow(x1, y1, x2, y2):
        draw.line((x1, y1, x2, y2), fill=BLACK, width=3)
        draw.polygon([(x2, y2), (x2 - 8, y2 - 12), (x2 + 8, y2 - 12)], fill=BLACK)

    # ===========================
    # PROCESSAMENTO
    # ===========================
    if not socios:
        draw.text((30, 30), "NO ENCONTRADO", fill=BLACK, font=f_big)
        return img

    socios = socios[:4]
    total = len(socios)

    total_width = total * box_w + (total - 1) * spacing
    start_x = (W - total_width) // 2
    top_y = 70

    centers = []

    # ===========================
    # DESENHA SCIOS
    # ===========================
    for i, s in enumerate(socios):
        x1 = start_x + i * (box_w + spacing)
        y1 = top_y
        x2 = x1 + box_w
        y2 = y1 + box_h

        # Caixa
        draw.rounded_rectangle([x1, y1, x2, y2], radius=20,
                               fill=BLUE, outline=BLUE_OUT, width=3)

        nome = (s.get("nome", "").strip().upper() or "")
        perc = (s.get("percentual", "").strip() or "")

        if perc and not perc.endswith("%"):
            perc += "%"

        # Quebra automtica
        nome_wrapped = wrap_text(nome, 16)

        # Fonte ideal
        font_sel = fit_font(nome, f_big, f_mid, f_small)

        draw_centered_multiline(x1, y1, x2, y2, nome_wrapped, font_sel)

        cx = (x1 + x2) // 2
        centers.append((cx, y2, perc))

    # ===========================
    # CAIXA DA EMPRESA
    # ===========================
    emp_x1 = (W - empresa_w) // 2
    emp_y1 = 280
    emp_x2 = emp_x1 + empresa_w
    emp_y2 = emp_y1 + empresa_h

    draw.rounded_rectangle([emp_x1, emp_y1, emp_x2, emp_y2],
                           radius=25, fill=BLUE, outline=BLUE_OUT, width=4)

    empresa_txt = abreviar(razao_social)
    empresa_txt = wrap_text(empresa_txt, 16)

    draw_centered_multiline(emp_x1, emp_y1, emp_x2, emp_y2, empresa_txt, f_big)

    # ===========================
    # LINHAS + SETAS
    # ===========================
    if total == 1:
        cx, bottom, perc = centers[0]
        mid_y = (bottom + emp_y1) // 2

        arrow(cx, bottom, cx, emp_y1)

        # percentual fora da linha ( esquerda)
        w, h = draw.textbbox((0, 0), perc, font=f_perc)[2:]
        draw.text((cx - 25 - w, mid_y - h / 2), perc, fill=BLACK, font=f_perc)

    else:
        # trilho curto
        left = centers[0][0]
        right = centers[-1][0]
        mid_y = 210
        emp_mid = (emp_x1 + emp_x2) // 2

        draw.line((left, mid_y, right, mid_y), fill=BLACK, width=3)

        for cx, bottom, perc in centers:
            draw.line((cx, bottom, cx, mid_y), fill=BLACK, width=3)

            py = (bottom + mid_y) // 2
            w, h = draw.textbbox((0, 0), perc, font=f_perc)[2:]
            draw.text((cx - 25 - w, py - h / 2), perc, fill=BLACK, font=f_perc)

        arrow(emp_mid, mid_y, emp_mid, emp_y1)

    return img

# ============================================================
#  EXTRATOR DE NOTÍCIAS — busca REAL no Google CSE (FILTRADO)
# ============================================================

import requests

# Domínios que NÃO são notícia (bases de dados / CNPJ)
DOMINIOS_BLOQUEADOS = [
    "casadosdados",
    "econodata",
    "cnpj",
    "cnpj.biz",
    "cnpjonline",
    "empresasdobrasil",
    "dadoscadastrais",
    "receitaws",
    "consultacnpj",
    "empresa.ninja"
]

def extract_noticias(razao_social: str) -> Dict[str, Any]:
    """
    Busca notícias reais sobre a empresa usando Google Custom Search.
    Filtra sites de base cadastral (CNPJ, dados empresariais).
    Retorna no formato exigido pelo Excel.
    """

    if not razao_social:
        return {"noticias": []}

    API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
    CX = os.getenv("GOOGLE_CSE_ID", "").strip()

    if not API_KEY or not CX:
        print("[ERRO] GOOGLE_API_KEY ou GOOGLE_CSE_ID ausentes no .env")
        return {"noticias": []}

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": API_KEY,
        "cx": CX,
        "q": f'"{razao_social}" notícias',
        "lr": "lang_pt",
        "num": 10
    }

    try:
        resp = requests.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()

        resultados = []

        for item in data.get("items", []):
            titulo = item.get("title", "") or ""
            resumo = item.get("snippet", "") or ""
            link = item.get("link", "") or ""

            # --- FILTRO DE DOMÍNIO (ponto-chave) ---
            link_low = link.lower()
            if any(dom in link_low for dom in DOMINIOS_BLOQUEADOS):
                continue

            # Fonte extraída do domínio
            fonte = ""
            if link:
                try:
                    fonte = link.split("/")[2]
                except:
                    fonte = ""

            # Data de publicação (se existir)
            data_pub = ""
            metatags = item.get("pagemap", {}).get("metatags", [{}])
            if isinstance(metatags, list) and metatags:
                pub = metatags[0].get("article:published_time")
                if pub:
                    data_pub = pub[:10]

            resultados.append({
                "titulo": titulo,
                "resumo": resumo,
                "fonte": fonte,
                "data": data_pub
            })

            # Limite real para o Excel
            if len(resultados) >= 4:
                break

        return {"noticias": resultados}

    except Exception as e:
        print("[ERRO extract_noticias]", e)
        return {"noticias": []}

# ============================================================
#  EXTRATOR DE CNPJ — busca REAL no Google CSE
# ============================================================

import re
import requests

def buscar_cnpj_por_razao(razao_social: str) -> str:
    """
    Usa Google Custom Search para localizar o CNPJ correspondente à razão social.
    Retorna o CNPJ no formato 00.000.000/0000-00 ou "" se não encontrar.
    """

    API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
    CX = os.getenv("GOOGLE_CSE_ID", "").strip()

    if not API_KEY or not CX:
        print("[ERRO] GOOGLE_API_KEY ou GOOGLE_CSE_ID ausentes no .env")
        return ""

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": API_KEY,
        "cx": CX,
        "q": f'"{razao_social}" CNPJ',
        "lr": "lang_pt",
        "num": 10
    }

    try:
        resp = requests.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()

        cnpj_regex = r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}"

        for item in data.get("items", []):
            texto = (item.get("title", "") + " " + item.get("snippet", ""))
            achado = re.findall(cnpj_regex, texto)
            if achado:
                print(f"[INFO] CNPJ encontrado: {achado[0]}")
                return achado[0]

        print("[AVISO] Nenhum CNPJ encontrado via Google.")
        return ""

    except Exception as e:
        print("[ERRO buscar_cnpj_por_razao]", e)
        return ""


# ============================================================
#  BUSCA DE ENDEREÇO REAL (CNPJ → Internet)
#  Agora muito mais robusto e seguro.
# ============================================================

import re
import requests

def limpar_endereco(end: str) -> str:
    """Remove lixo do Google e deixa o endereço limpo."""
    if not end:
        return ""
    end = end.replace("\n", " ").strip()
    end = re.sub(r"\s+", " ", end)
    return end


def validar_endereco(end: str) -> bool:
    if not end:
        return False

    end_low = end.lower()

    tem_rua = any(x in end_low for x in [
        "rua", "avenida", "av ", "travessa", "rodovia", "alameda", "estrada", "r."
    ])

    tem_numero = bool(re.search(r",\s*\d{1,5}", end))
    tem_uf = bool(re.search(r"\b(ac|al|ap|am|ba|ce|df|es|go|ma|mt|ms|mg|pa|pb|pr|pe|pi|rj|rn|rs|ro|rr|sc|sp|se|to)\b", end_low))
    tem_virgulas = end.count(",") >= 2

    return tem_rua and tem_numero and tem_uf and tem_virgulas


def buscar_endereco_google(query: str) -> str:
    """Executa a busca real no Google CSE."""
    API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
    CX = os.getenv("GOOGLE_CSE_ID", "").strip()

    if not API_KEY or not CX:
        print("[ERRO] GOOGLE_API_KEY ou GOOGLE_CSE_ID ausentes no .env")
        return ""

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": API_KEY,
        "cx": CX,
        "q": query,
        "lr": "lang_pt",
        "num": 10
    }

    try:
        resp = requests.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()

        endereco_regex = r"(Rua|Avenida|Av\.?|Travessa|Rodovia|Alameda|Estrada)[^,]{3,100},[^,]{3,100},[^,]{3,100}"

        for item in data.get("items", []):
            texto = (item.get("title", "") + " " + item.get("snippet", ""))
            achados = re.findall(endereco_regex, texto, re.IGNORECASE)

            if achados:
                end = limpar_endereco(achados[0])
                if validar_endereco(end):
                    print(f"[INFO] Endereço real encontrado: {end}")
                    return end

        return ""

    except Exception as e:
        print("[ERRO buscar_endereco_google]", e)
        return ""


def buscar_endereco_por_cnpj(cnpj: str, razao: str) -> str:
    """
    Tenta extrair o endereço REAL usando Google Search (CSE),
    com heurísticas agressivas para encontrar e validar texto de endereço.
    """

    API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
    CX = os.getenv("GOOGLE_CSE_ID", "").strip()

    if not API_KEY or not CX:
        print("[ERRO] GOOGLE_API_KEY ou GOOGLE_CSE_ID ausentes no .env")
        return ""

    if not cnpj:
        print("[AVISO] CNPJ vazio — tentativa via razão social.")
    
    import re
    candidatos = []

    def fazer_busca(query: str):
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            "key": API_KEY,
            "cx": CX,
            "q": query,
            "lr": "lang_pt",
            "num": 10
        }

        try:
            resp = requests.get(url, params=params)
            resp.raise_for_status()
            return resp.json().get("items", [])
        except:
            return []

    # ==========================================================
    # 1) PRIMEIRA TENTATIVA — CNPJ
    # ==========================================================
    if cnpj:
        print(f"[INFO] Buscando endereço REAL via CNPJ {cnpj}...")
        resultados = fazer_busca(f'"{cnpj}" endereço sede matriz localizacao')
        candidatos.extend(resultados)

    # ==========================================================
    # 2) SEGUNDA TENTATIVA — Razão Social
    # ==========================================================
    if razao:
        print(f"[INFO] Buscando endereço REAL via razão social '{razao}'...")
        resultados = fazer_busca(f'"{razao}" endereço sede matriz CNPJ endereço completo')
        candidatos.extend(resultados)

    if not candidatos:
        return ""

    # ==========================================================
    # 3) REGEX ROBUSTA DE ENDEREÇOS
    # - Rua / Av / Alameda / Rodovia / Travessa
    # - Número obrigatório
    # - Cidade + UF obrigatórios
    # ==========================================================

    padrao_endereco = re.compile(
        r"(Rua|Avenida|Av\.?|Travessa|Alameda|Rodovia|Estrada|R\.)\s+"
        r"[A-Za-zÀ-ÿ0-9\s\.]{3,60},\s*"
        r"\d{1,5}"
        r"(?:\s*-\s*[A-Za-zÀ-ÿ\s]{2,40})?,\s*"
        r"[A-Za-zÀ-ÿ\s]{3,40}\s*-\s*"
        r"(AC|AL|AP|AM|BA|CE|DF|ES|GO|MA|MT|MS|MG|PA|PB|PR|PE|PI|RJ|RN|RS|RO|RR|SC|SP|SE|TO)",
        flags=re.IGNORECASE
    )

    def score_endereco(e: str) -> int:
        score = 0
        if re.search(r",\s*\d{1,5}", e):
            score += 3
        if e.count(",") >= 2:
            score += 2
        if "-" in e:
            score += 1
        return score

    melhores = []

    for item in candidatos:
        texto = (item.get("title", "") + " " + item.get("snippet", "")).strip()

        match = padrao_endereco.search(texto)
        if match:
            end = match.group(0).strip()
            melhores.append(end)

    if not melhores:
        print("[AVISO] Nenhum endereço com padrão reconhecido.")
        return ""

    # Escolhe o endereço mais bem estruturado
    melhor = max(melhores, key=score_endereco)
    melhor = melhor.strip().rstrip(".,;:-")

    return melhor


# ============================================================
#  EXTRATOR DE JURÍDICO — busca REAL no Google CSE
# ============================================================

def extract_juridico(razao_social: str, socios: List[str], cnpj: str) -> Dict[str, Any]:
    """
    Busca jurídica real usando Google Search:
    - Para cada sócio (nome)
    - Para a empresa (CNPJ)
    
    Retorna:
    {
        "empresa_info": [...],
        "socio_juridico": [...]
    }
    """

    API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
    CX = os.getenv("GOOGLE_CSE_ID", "").strip()

    if not API_KEY or not CX:
        print("[ERRO] GOOGLE_API_KEY ou GOOGLE_CSE_ID ausentes no .env")
        return {"empresa_info": [], "socio_juridico": []}

    def google_busca(query: str):
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            "key": API_KEY,
            "cx": CX,
            "q": query,
            "lr": "lang_pt",
            "num": 10
        }

        resultados = []
        try:
            resp = requests.get(url, params=params)
            resp.raise_for_status()
            data = resp.json()

            for item in data.get("items", []):
                titulo = item.get("title", "")
                snippet = item.get("snippet", "")
                link = item.get("link", "")
                resultados.append(f"{titulo} — {snippet} ({link})")

        except Exception as e:
            print("[ERRO google_busca]", e)

        return resultados if resultados else ["Nenhuma informação encontrada."]

    # ---------------------------
    # BUSCA DO SÓCIOS
    # ---------------------------
    socio_juridico = []
    for socio in socios:
        query = (
            f'"{socio}" processo OR ação OR judicial OR execução '
            f'OR trabalhista OR jusbrasil OR escavador'
        )
        socio_juridico.append({
            "socio": socio,
            "informacoes": google_busca(query)
        })

    # ---------------------------
    # BUSCA DO CNPJ DA EMPRESA
    # ---------------------------
    empresa_info = []
    if cnpj:
        query_empresa = (
            f'"{cnpj}" processo OR ação OR judicial OR execução '
            f'OR trabalhista OR jusbrasil OR escavador'
        )
        empresa_info = google_busca(query_empresa)

    return {
        "empresa_info": empresa_info,
        "socio_juridico": socio_juridico
    }


# ------------------------------------------------------------
# MAPA DE NORMALIZAO DE MESES (corrige abreviacoes do GPT)
# ------------------------------------------------------------

import unicodedata

MAP_MESES = {
    "jan": "janeiro",
    "janeiro": "janeiro",
    "1": "janeiro",
    "01": "janeiro",

    "fev": "fevereiro",
    "fevereiro": "fevereiro",
    "2": "fevereiro",
    "02": "fevereiro",

    "mar": "marco",
    "marc": "marco",
    "marco": "marco",
    "marco.": "marco",
    "maro": "marco",
    "marco": "marco",  # bug unicode
    "3": "marco",
    "03": "marco",

    "abr": "abril",
    "abril": "abril",
    "4": "abril",
    "04": "abril",

    "mai": "maio",
    "maio": "maio",
    "5": "maio",
    "05": "maio",

    "jun": "junho",
    "junho": "junho",
    "6": "junho",
    "06": "junho",

    "jul": "julho",
    "julho": "julho",
    "7": "julho",
    "07": "julho",

    "ago": "agosto",
    "agosto": "agosto",
    "8": "agosto",
    "08": "agosto",

    "set": "setembro",
    "setembro": "setembro",
    "9": "setembro",
    "09": "setembro",

    "out": "outubro",
    "outubro": "outubro",
    "10": "outubro",

    "nov": "novembro",
    "novembro": "novembro",
    "11": "novembro",

    "dez": "dezembro",
    "dezembro": "dezembro",
    "12": "dezembro",
}


def normalizar_mes(mes_raw: str) -> str:
    """
    Normaliza qualquer forma de ms:
    - remove acentos
    - remove caracteres estranhos do OCR
    - entende abreviaes (jan, fev, mar)
    - aceita nmeros (0112)
    - limpa formatos tipo 'mar/24', 'mar-2025', 'mar25'
    """

    if not isinstance(mes_raw, str):
        return ""

    # Normalizao Unicode (remove acentos e caracteres invisveis)
    mes = unicodedata.normalize("NFKD", mes_raw)
    mes = "".join(c for c in mes if not unicodedata.combining(c))
    mes = mes.lower().strip()

    # Remove ano grudado -> "mar25", "mar/24"
    for sep in ["/", "-", "_", " "]:
        if sep in mes:
            mes = mes.split(sep)[0]

    # Se terminou com nmero (ex: "mar25") -> remove nmeros finais
    while mes and mes[-1].isdigit():
        mes = mes[:-1]

    mes = mes.strip()

    # Busca direta
    if mes in MAP_MESES:
        return MAP_MESES[mes]

    # Busca por prefixo (ex: "marc", "marc.", "mar")
    for k in MAP_MESES:
        if mes.startswith(k):
            return MAP_MESES[k]

    return mes  # fallback

# ============================================================
#  BUSCA DE ENDEREÇO VIA GOOGLE PLACES (fallback)
# ============================================================

import requests

def buscar_endereco_por_nome(nome_empresa: str, cidade: str = ""):
    """
    Busca endereço e coordenadas pelo nome da empresa usando Google Places.
    Retorna: (endereco, lat, lng)
    """

    API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()

    if not API_KEY or not nome_empresa:
        print("[AVISO] API do Maps ausente ou nome da empresa vazio.")
        return "", None, None

    query = nome_empresa
    if cidade:
        query += f", {cidade}"

    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {
        "query": query,
        "language": "pt-BR",
        "key": API_KEY
    }

    try:
        resp = requests.get(url, params=params)
        data = resp.json()

        if data.get("status") != "OK":
            print(f"[AVISO] Places retornou: {data.get('status')}")
            return "", None, None

        result = data["results"][0]

        endereco = result.get("formatted_address", "")
        location = result.get("geometry", {}).get("location", {})

        lat = location.get("lat")
        lng = location.get("lng")

        print(f"[INFO] Endereço encontrado via Places: {endereco}")
        return endereco, lat, lng

    except Exception as e:
        print(f"[ERRO] buscar_endereco_por_nome falhou: {e}")
        return "", None, None

# ============================================================
#  GOOGLE MAPS API — Busca de endereço, satélite e street view
# ============================================================

def maps_ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


def maps_get_coordinates(endereco: str):
    """Converte endereço para latitude e longitude usando Google Geocoding API."""
    API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()

    if not API_KEY:
        print("[ERRO] GOOGLE_MAPS_API_KEY ausente no .env")
        return None, None

    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": endereco, "key": API_KEY}

    try:
        data = requests.get(url, params=params).json()

        if data.get("status") != "OK":
            print("[AVISO] Geocoding não retornou resultado:", data.get("status"))
            return None, None

        # 🔹 Prioriza endereço real (imóvel)
        for r in data.get("results", []):
            types = r.get("types", [])
            if "street_address" in types or "premise" in types:
                loc = r["geometry"]["location"]
                return loc["lat"], loc["lng"]

        # 🔹 Fallback
        loc = data["results"][0]["geometry"]["location"]
        return loc["lat"], loc["lng"]

    except Exception as e:
        print("[ERRO] maps_get_coordinates:", e)
        return None, None



def maps_download_satellite(lat, lng, output_path: str, zoom=20, size="640x640"):
    """Baixa imagem de satélite usando Static Maps API."""
    API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()

    url = "https://maps.googleapis.com/maps/api/staticmap"
    params = {
        "center": f"{lat},{lng}",
        "zoom": zoom,
        "size": size,
        "maptype": "satellite",
        "key": API_KEY
    }

    try:
        resp = requests.get(url, params=params)
        if resp.status_code != 200:
            raise Exception("Erro HTTP")
        with open(output_path, "wb") as f:
            f.write(resp.content)
        return output_path
    except Exception as e:
        print("[ERRO] maps_download_satellite:", e)
        return ""


def maps_download_streetview(lat, lng, output_path, size="600x600"):
    API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()

    url = "https://maps.googleapis.com/maps/api/streetview"
    params = {
        "location": f"{lat},{lng}",
        "size": size,
        "source": "outdoor",
        "radius": 50,
        "pitch": 0,
        "fov": 90,
        "key": API_KEY
    }

    try:
        resp = requests.get(url, params=params)
        if resp.status_code != 200:
            raise Exception("Erro HTTP")
        with open(output_path, "wb") as f:
            f.write(resp.content)
        return output_path
    except Exception as e:
        print("[ERRO] maps_download_streetview:", e)
        return ""



def fetch_maps_images(
    endereco: str,
    save_dir: str,
    lat_override=None,
    lng_override=None
):
    """
    Endereço → coordenadas → imagens (satélite + street view).
    Permite override de coordenadas (ex: Google Places).
    """
    maps_ensure_dir(save_dir)

    if lat_override is not None and lng_override is not None:
        lat, lng = lat_override, lng_override
        print("[INFO] Usando coordenadas fornecidas diretamente.")
    else:
        print(f"📍 Buscando coordenadas para: {endereco}")
        lat, lng = maps_get_coordinates(endereco)

    if not lat or not lng:
        print("[AVISO] Coordenadas não encontradas, módulo Maps desativado.")
        return {
            "latitude": None,
            "longitude": None,
            "satellite_image": "",
            "streetview_image": ""
        }

    sat_path = os.path.join(save_dir, "satellite.jpg")
    street_path = os.path.join(save_dir, "streetview.jpg")

    print("🛰 Baixando imagem de satélite...")
    sat_file = maps_download_satellite(lat, lng, sat_path)

    print("📸 Baixando imagem do Street View...")
    street_file = maps_download_streetview(lat, lng, street_path)

    print("✅ Imagens de mapa obtidas com sucesso!")

    return {
        "latitude": lat,
        "longitude": lng,
        "satellite_image": sat_file,
        "streetview_image": street_file
    }

# ============================================================
# 10. MONTAGEM DO EXCEL FINAL (VERSO ESTAVEL E AJUSTADA)
# ============================================================

def preencher_excel(
    modelo_path: str,
    destino_path: str,
    contrato: Dict[str, Any],
    faturamentos: Dict[str, Dict[str, Any]],
    endivid_bancos: List[Dict[str, Any]],
    endivid_fundos: List[Dict[str, Any]],
    juridico: Dict[str, Any],
    noticias: Dict[str, Any]
):

    from openpyxl.styles import Font

    # 1  Copiar modelo
    try:
        shutil.copy(modelo_path, destino_path)
    except Exception as e:
        print(f"[ERRO] Falha ao copiar modelo: {e}")
        return

    wb = load_workbook(destino_path)

    # --------------------------------------------------------
    # CAPA
    # --------------------------------------------------------
    try:
        ws = wb["CAPA"]
        ws["C26"] = contrato.get("razao_social", "")
    except:
        print("[AVISO] Aba CAPA no encontrada.")

    # --------------------------------------------------------
    # DADOS EMPRESA
    # --------------------------------------------------------
    try:
        ws = wb["DADOS EMPRESA"]

        cap = to_number_br(contrato.get("capital_social", ""))
        ws["B7"] = cap

        ws["C8"] = contrato.get("cidade", "")
        ws["D8"] = contrato.get("estado", "")

    except:
        print("[AVISO] Aba DADOS EMPRESA no encontrada.")

    # --------------------------------------------------------
    # ESTRUTURA SOCIETÁRIA
    # --------------------------------------------------------
    try:
        ws = wb["ESTRUTURA SOCIETÁRIA"]
        socios = contrato.get("socios", [])

        if not socios:
            ws["A6"] = "NO ENCONTRADO"

        for i in range(4):
            if i < len(socios):
                ws[f"A{6+i}"] = socios[i].get("nome", "")
                ws[f"B{6+i}"] = socios[i].get("percentual", "")
            else:
                ws[f"A{6+i}"] = ""
                ws[f"B{6+i}"] = ""

    except Exception as e:
        print(f"[AVISO] Aba ESTRUTURA SOCIETÁRIA falhou: {e}")

    # --------------------------------------------------------
    # ORGANOGRAMA
    # --------------------------------------------------------
    try:
        ws = wb["ORGANOGRAMA"]
        img = build_organograma(
            contrato.get("socios", []),
            contrato.get("razao_social", "")
        )
        insert_image(ws, img, "A5")
    except Exception as e:
        print(f"[AVISO] Organograma falhou: {e}")


    # --------------------------------------------------------
    # FACHADA (Street View)
    # --------------------------------------------------------
    try:
        ws = wb["FACHADA"]

        img_path = contrato.get("maps_street", "")

        if img_path and os.path.isfile(img_path):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(img_path)

            # opcional — ajustar tamanho aproximado para caber em A5:D28
            img.width = 650
            img.height = 500

            ws.add_image(img, "A5")
        else:
            ws["A5"] = "Imagem de fachada não encontrada."

    except Exception as e:
        print("[AVISO] Aba FACHADA falhou:", e)


    # --------------------------------------------------------
    # SATÉLITE
    # --------------------------------------------------------
    try:
        ws = wb["SATELITE"]

        img_path = contrato.get("maps_satellite", "")

        if img_path and os.path.isfile(img_path):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(img_path)

            # Ajuste de tamanho
            img.width = 650
            img.height = 500

            ws.add_image(img, "A5")
        else:
            ws["A5"] = "Imagem de satélite não encontrada."

    except Exception as e:
        print("[AVISO] Aba SATELITE falhou:", e)

    # --------------------------------------------------------
    # FATURAMENTO (como NMEROS)
    # --------------------------------------------------------
    try:
        ws = wb["FAT"]

        map_ano_col = {"2025": "B", "2024": "E", "2023": "H"}

        ordem_meses = [
            "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
            "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
        ]

        for ano, col in map_ano_col.items():
            meses = faturamentos.get(ano, {}).get("meses", {})
            for i, mes in enumerate(ordem_meses):
                num = to_number_br(meses.get(mes, ""))
                ws[f"{col}{6+i}"] = num

    except Exception as e:
        print(f"[AVISO] Faturamento falhou: {e}")

    # --------------------------------------------------------
    # INSTITUIÇÕES (Bancos + Fundos unificados)
    # --------------------------------------------------------
    try:
        ws = wb["INSTITUIÇÕES"]

        row = 7
        max_rows = 26

        todas = endivid_bancos + endivid_fundos

        if not todas:
            ws["A7"] = "NADA ENCONTRADO"
        else:
            for item in todas:
                if row > max_rows:
                    break

                credor = item.get("credor", "").strip()
                modalidade = item.get("modalidade", "").strip()
                saldo = to_number_br(item.get("saldo_devedor", ""))

                # Regra de classificação
                if eh_banco(credor):
                    modalidade_final = modalidade if modalidade else "CRÉDITO"
                else:
                    modalidade_final = "DESCONTO"

                ws[f"A{row}"] = credor
                ws[f"B{row}"] = modalidade_final
                ws[f"C{row}"] = saldo

                row += 1

    except Exception as e:
        print(f"[AVISO] Aba INSTITUICOES falhou: {e}")


    # --------------------------------------------------------
    # JURDICO IA (EMPRESA + SCIOS)  Fonte 12
    # --------------------------------------------------------
    try:
        ws = wb["JURIDICO IA"]

        row = 5
        socio_info = juridico.get("socio_juridico", [])
        nome_empresa = contrato.get("razao_social", "").strip()

        # --- EMPRESA ---
        if nome_empresa:
            ws[f"A{row}"] = f"Empresa: {nome_empresa}"
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 1

            if juridico.get("empresa_info"):
                for info in juridico["empresa_info"]:
                    ws[f"A{row}"] = f" {info}"
                    ws[f"A{row}"].font = Font(size=12)
                    row += 1
            else:
                ws[f"A{row}"] = " Nada encontrado"
                ws[f"A{row}"].font = Font(size=12)
                row += 1

            row += 1

        # --- SOCIOS ---
        if socio_info:
            for item in socio_info:
                socio = item.get("socio", "")
                ws[f"A{row}"] = f"Scio: {socio}"
                ws[f"A{row}"].font = Font(size=12, bold=True)
                row += 1

                detalhes = item.get("informacoes", [])
                if detalhes:
                    for info in detalhes:
                        ws[f"A{row}"] = f" {info}"
                        ws[f"A{row}"].font = Font(size=12)
                        row += 1
                else:
                    ws[f"A{row}"] = " Nada encontrado"
                    ws[f"A{row}"].font = Font(size=12)
                    row += 1

                row += 1
        else:
            ws[f"A{row}"] = "Nenhuma informao jurdica encontrada."
            ws[f"A{row}"].font = Font(size=12)

    except Exception as e:
        print(f"[AVISO] Aba JURIDICO IA falhou: {e}")

    # --------------------------------------------------------
    # NOTICIAS  TItulo, resumo, fonte, data
    # --------------------------------------------------------
    try:
        ws = wb["NOTICIAS"]

        lista = noticias.get("noticias", [])
        row = 5

        if not lista:
            ws["A5"] = "Nenhuma notcia encontrada."
        else:
            for n in lista[:4]:
                titulo = n.get("titulo", "")
                resumo = n.get("resumo", "")
                fonte = n.get("fonte", "")
                data = n.get("data", "")

                ws[f"A{row}"] = titulo
                ws[f"A{row}"].font = Font(size=12, bold=True)
                row += 1

                ws[f"A{row}"] = resumo
                ws[f"A{row}"].font = Font(size=12)
                row += 1

                ws[f"A{row}"] = f"Fonte: {fonte}  {data}"
                ws[f"A{row}"].font = Font(size=12, italic=True)
                row += 2

    except Exception as e:
        print(f"[AVISO] Aba NOTICIAS falhou: {e}")

# --------------------------------------------------------
# PRESERVAR PAGINAÇÃO / LOGO DO MODELO
# --------------------------------------------------------

    try:
        for ws in wb.worksheets:
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = False
    except:
        pass

# --------------------------------------------------------
# SALVAR
# --------------------------------------------------------

    try:
        wb.save(destino_path)
        print(f"[INFO] Excel gerado com sucesso: {destino_path}")
    except Exception as e:
        print(f"[ERRO] Falha ao salvar Excel: {e}")

# ============================================================
# IDENTIFICAÇÃO DA EMPRESA PELO PREFIXO DO ARQUIVO
# ============================================================

import re

def identificar_empresa_por_arquivo(nome_arquivo: str) -> str:
    """
    Identifica a empresa pelo prefixo do nome do arquivo.

    Exemplos aceitos:
      01. Contrato Social.pdf  -> "01"
      02. Faturamento 2024.pdf -> "02"
      03.Endividamento.pdf    -> "03"

    Retorna:
      "01", "02", "03" ...
      ou "" se não identificar
    """

    if not nome_arquivo:
        return ""

    nome = nome_arquivo.strip()

    match = re.match(r"^(\d{2})\s*\.", nome)
    if match:
        return match.group(1)

    return ""

# ============================================================
# 11. LOCALIZAO AUTOMTICA DOS PDFs (VERSO ESTAVEL)
# ============================================================

def localizar_pdfs(pasta_cliente: str) -> Dict[str, Any]:
    """
    Localiza automaticamente:
      Contrato Social  -> pasta 06. PJ
      Faturamento      -> pasta 02. FATURAMENTO
      Endividamento    -> pasta 03. ENDIVIDAMENTO
    """

    pasta_contrato = os.path.join(pasta_cliente, "06. PJ")
    pasta_faturamento = os.path.join(pasta_cliente, "02. FATURAMENTO")
    pasta_endivid = os.path.join(pasta_cliente, "03. ENDIVIDAMENTO")

    contrato_pdf = None
    faturamento_pdfs = []
    endivid_pdfs = []

    # -------- CONTRATO SOCIAL --------
    if os.path.isdir(pasta_contrato):
        for f in os.listdir(pasta_contrato):
            fn = f.lower()
            if fn.endswith(".pdf") and ("contrato" in fn or "social" in fn):
                contrato_pdf = os.path.join(pasta_contrato, f)
                break

    # -------- FATURAMENTO --------
    if os.path.isdir(pasta_faturamento):
        for f in os.listdir(pasta_faturamento):
            fn = f.lower()

            if not fn.endswith(".pdf"):
                continue

            caminho = os.path.join(pasta_faturamento, f)

            # Se conter ano -> prioriza
            if any(ano in fn for ano in ["2023", "2024", "2025"]):
                faturamento_pdfs.append(caminho)
                continue

            # Caso misto -> tambm adiciona
            faturamento_pdfs.append(caminho)

    # -------- ENDIVIDAMENTO --------
    if os.path.isdir(pasta_endivid):
        for f in os.listdir(pasta_endivid):
            fn = f.lower()
            if fn.endswith(".pdf") and ("endivid" in fn or "divida" in fn):
                endivid_pdfs.append(os.path.join(pasta_endivid, f))

    return {
        "contrato": contrato_pdf,
        "faturamento": faturamento_pdfs,
        "endividamento": endivid_pdfs
    }

# ============================================================
# 12. FLUXO PRINCIPAL (MAIN)  VERSO ESTAVEL 2025
# ============================================================

def main():
    print("========================================================")
    print("     BOT DE ANLISE DE CRDITO - ULTRA HBRIDO (v3)")
    print("========================================================\n")

    # --------------------------------------------------------
    # CARREGA LISTA EXTERNA DE BANCOS
    # --------------------------------------------------------
    carregar_bancos()

    # --------------------------------------------------------
    # Recebe caminho da pasta do cliente
    # --------------------------------------------------------
    if len(sys.argv) < 2:
        print("[ERRO] Caminho da pasta do cliente no informado.")
        print('Uso correto: bot.exe "C:\\pasta\\cliente"')
        sys.exit(1)

    pasta_cliente = sys.argv[1].strip('"')
    print("Pasta do cliente:\n" + pasta_cliente + "\n")

    if not os.path.isdir(pasta_cliente):
        print("[ERRO] Caminho informado no  uma pasta vlida.")
        sys.exit(1)

    # --------------------------------------------------------
    # Localiza automaticamente todos os PDFs
    # --------------------------------------------------------
    print("--------------------------------------------------------")
    print("Localizando PDFs...")
    arquivos = localizar_pdfs(pasta_cliente)
    print("Contrato:", arquivos["contrato"])
    print("Faturamento:", arquivos["faturamento"])
    print("Endividamento:", arquivos["endividamento"])
    print("--------------------------------------------------------\n")

    # --------------------------------------------------------
    # CONTRATO SOCIAL
    # --------------------------------------------------------
    print("Extraindo CONTRATO SOCIAL...")

    cnpj_empresa = ""
    contrato = {
        "razao_social": "",
        "cidade": "",
        "estado": "",
        "capital_social": "",
        "socios": []
    }

    if arquivos["contrato"]:
        extraido = extract_contrato(arquivos["contrato"])

        # substitui somente se o extrator retornou algo válido
        if extraido:
            contrato = extraido

            # tenta buscar CNPJ
            if contrato.get("razao_social"):
                cnpj_empresa = buscar_cnpj_por_razao(contrato["razao_social"])
    else:
        print("[AVISO] Nenhum contrato social encontrado.")


    # --------------------------------------------------------
    # MAPS — buscar endereço real pelo CNPJ e baixar imagens
    # --------------------------------------------------------
    print("\nConsultando endereço real da empresa...")

    # 1) Tenta pelo CNPJ
    endereco_empresa = buscar_endereco_por_cnpj(
        cnpj_empresa,
        contrato.get("razao_social", "")
    )

    # 2) Se falhar, tenta pelo nome da empresa (Google Places)
    if not endereco_empresa:
        print("[INFO] Tentando obter endereço pelo nome da empresa...")
        endereco_empresa, lat_force, lng_force = buscar_endereco_por_nome(
            contrato.get("razao_social", ""),
            contrato.get("cidade", "")
        )

        if endereco_empresa and lat_force and lng_force:
            print("[INFO] Coordenadas obtidas via Google Places.")

            maps_result = fetch_maps_images(
                endereco_empresa,
                save_dir=os.path.join(pasta_cliente, "MAPS"),
                lat_override=lat_force,
                lng_override=lng_force
            )

            contrato["maps_satellite"] = maps_result["satellite_image"]
            contrato["maps_street"] = maps_result["streetview_image"]
        else:
            print("[AVISO] Nenhum endereço confirmado — ignorando Maps.")
            contrato["maps_satellite"] = ""
            contrato["maps_street"] = ""
    else:
        # Endereço via CNPJ funcionou → segue o processo normal
        maps_result = fetch_maps_images(
            endereco_empresa,
            save_dir=os.path.join(pasta_cliente, "MAPS")
        )
        contrato["maps_satellite"] = maps_result["satellite_image"]
        contrato["maps_street"] = maps_result["streetview_image"]


    # --------------------------------------------------------
    # FATURAMENTO
    # --------------------------------------------------------
    print("Extraindo FATURAMENTO...")

    faturamentos = {
        "2023": {"meses": {}},
        "2024": {"meses": {}},
        "2025": {"meses": {}}
    }

    for pdf in arquivos["faturamento"]:
        print(f"  Analisando: {os.path.basename(pdf)}")
        dados = extract_faturamento(pdf)

        for item in dados.get("dados", []):
            ano = str(item.get("ano", "")).strip()
            mes = normalizar_mes(item.get("mes", ""))
            valor = item.get("valor", "")

            if ano in faturamentos and mes:
                faturamentos[ano]["meses"][mes] = valor


    # --------------------------------------------------------
    # ENDIVIDAMENTO — unificado com lista externa de bancos
    # --------------------------------------------------------
    print("Extraindo ENDIVIDAMENTO...")

    endivid_bancos = []
    endivid_fundos = []

    for pdf in arquivos["endividamento"]:
        itens = extract_endividamento(pdf)

    for it in itens:
        credor_nome = it.get("credor", "").strip()

        # Classificação REAL usando bancos.txt
        if eh_banco(credor_nome):
            it["tipo_credor"] = "banco"
            endivid_bancos.append(it)
        else:
            it["tipo_credor"] = "desconto"
            endivid_fundos.append(it)


    # --------------------------------------------------------
    # JURÍDICO (Google CSE – empresa + sócios)
    # --------------------------------------------------------
    print("Extraindo JURIDICO REAL...\n")

    # Lista de nomes dos sócios extraídos do contrato
    lista_socios = [
        s.get("nome", "").strip()
        for s in contrato.get("socios", [])
        if s.get("nome")
    ]

    print(f"[INFO] Sócios identificados: {lista_socios if lista_socios else 'Nenhum sócio encontrado'}")
    print(f"[INFO] CNPJ identificado para pesquisa jurídica: {cnpj_empresa if cnpj_empresa else 'Não encontrado'}")
    print("\nPesquisando informações jurídicas reais...\n")

    juridico = extract_juridico(
        razao_social = contrato.get("razao_social", ""),
        socios       = lista_socios,
        cnpj         = cnpj_empresa
    )

    # Feedback de conclusão no console
    qt_soc = len(juridico.get("socio_juridico", []))
    qt_emp = len(juridico.get("empresa_info", []))

    print(f"[INFO] Busca jurídica concluída.")
    print(f"       → Informações da empresa: {qt_emp} registros")
    print(f"       → Informações dos sócios: {qt_soc} sócios analisados\n")


    # --------------------------------------------------------
    # NOTCIAS
    # --------------------------------------------------------
    print("Extraindo NOTCIAS...")
    noticias = extract_noticias(contrato.get("razao_social", ""))

    # --------------------------------------------------------
    # Montagem do Excel final
    # --------------------------------------------------------
    print("\nGerando anlise final...")

    modelo_path = os.path.join(EXEC_DIR, "novo modelo de analise.xlsx")
    destino_path = os.path.join(pasta_cliente, "ANALISE FINAL.xlsx")

    preencher_excel(
        modelo_path=modelo_path,
        destino_path=destino_path,
        contrato=contrato,
        faturamentos=faturamentos,
        endivid_bancos=endivid_bancos,
        endivid_fundos=endivid_fundos,
        juridico=juridico,
        noticias=noticias
    )

    print("========================================================")
    print("ANLISE FINALIZADA  NED CAPITAL")
    print("Arquivo gerado:")
    print(destino_path)
    print("========================================================")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("\n============================================")
        print("     ERRO FATAL NA EXECUCAO DO BOT")
        print("============================================")
        print(e)
        traceback.print_exc()
        print("============================================\n")
        print("Aperte ENTER para fechar...")
        try:
            input()
        except:
            os.system("pause")

